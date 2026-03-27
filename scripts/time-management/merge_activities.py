"""
Merge consecutive Timing export rows that share the same non-empty Path and are
adjacent in time (gap <= tolerance). Empty Path rows are never merged.

The original workbook is NEVER modified. Merged output is written to a separate
file: ``Merged Activities.xlsx`` (same directory as the source).

Optional: require same Application as well by passing require_same_application=True
if you see incorrect merges across apps.
"""

from __future__ import annotations

import os
import sys
from urllib.parse import urlparse, urlunparse

import pandas as pd
from openpyxl import load_workbook

DEFAULT_GAP_SECONDS = 2.0
MERGED_FILE_NAME = "Merged Activities.xlsx"
# Excel stores durations as fraction of a day; without this format, users see decimals.
EXCEL_DURATION_FORMAT = "[h]:mm:ss"


def normalize_path(value) -> str | None:
    """Return the raw path string, or None when empty/missing."""
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    s = str(value).strip()
    if not s or s.lower() == "none":
        return None
    return s


def clean_url(raw: str) -> str:
    """
    Produce a canonical merge key from a URL by stripping parts that don't
    change which page is being viewed:
      - query string  (?q=...&tracking=...)
      - fragment      (#section)
      - trailing /
      - www. prefix   (www.linkedin.com -> linkedin.com)

    Non-URL paths (file paths, empty strings) are returned as-is.
    """
    if not raw.startswith(("http://", "https://")):
        return raw

    parsed = urlparse(raw)
    host = parsed.netloc
    if host.startswith("www."):
        host = host[4:]

    path = parsed.path.rstrip("/")
    return urlunparse((parsed.scheme, host, path, "", "", ""))


def merge_consecutive_activities(
    df: pd.DataFrame,
    gap_seconds: float = DEFAULT_GAP_SECONDS,
    require_same_application: bool = False,
) -> pd.DataFrame:
    if df.empty:
        return df.copy()

    out_cols = [c for c in df.columns]
    work = df.copy()
    work["_start"] = pd.to_datetime(work["Start Date"], errors="coerce")
    work["_end"] = pd.to_datetime(work["End Date"], errors="coerce")

    work = work.sort_values("_start", na_position="last").reset_index(drop=True)

    n = len(work)
    results: list[dict] = []
    i = 0

    while i < n:
        row = work.iloc[i]
        start = row["_start"]
        end = row["_end"]
        raw_path = normalize_path(row.get("Path"))

        if raw_path is None or pd.isna(start) or pd.isna(end):
            results.append(_row_dict_from_slice(row, out_cols, start, end))
            i += 1
            continue

        path_key = clean_url(raw_path)
        run_start = start
        run_end = end
        base_app = row.get("Application")
        j = i + 1

        while j < n:
            nxt = work.iloc[j]
            nxt_raw = normalize_path(nxt.get("Path"))
            if nxt_raw is None:
                break
            if clean_url(nxt_raw) != path_key:
                break
            if require_same_application and nxt.get("Application") != base_app:
                break

            ns = nxt["_start"]
            ne = nxt["_end"]
            if pd.isna(ns) or pd.isna(ne):
                break

            delta = (ns - run_end).total_seconds()
            if delta > gap_seconds:
                break

            run_end = max(run_end, ne)
            j += 1

        first = work.iloc[i]
        merged = {c: first[c] for c in out_cols}
        merged["Start Date"] = run_start
        merged["End Date"] = run_end
        if "Duration" in out_cols:
            merged["Duration"] = run_end - run_start
        if "Day" in out_cols:
            merged["Day"] = run_start.floor("D")
        results.append(merged)
        i = j

    return pd.DataFrame(results, columns=out_cols)


def _row_dict_from_slice(row, out_cols, start, end) -> dict:
    out = {c: row[c] for c in out_cols}
    if pd.notna(start) and pd.notna(end) and "Duration" in out_cols:
        out["Duration"] = end - start
    if "Day" in out_cols and pd.notna(start):
        out["Day"] = start.floor("D")
    return out


def _format_duration_column(xlsx_path: str, sheet_name: str, column_name: str = "Duration") -> None:
    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return
    ws = wb[sheet_name]
    header = [cell.value for cell in ws[1]]
    try:
        col_idx = header.index(column_name) + 1
    except ValueError:
        wb.close()
        return
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None:
            cell.number_format = EXCEL_DURATION_FORMAT
    wb.save(xlsx_path)
    wb.close()


def write_merged_file(
    xlsx_path: str,
    source_sheet: str | int = 0,
    gap_seconds: float = DEFAULT_GAP_SECONDS,
    require_same_application: bool = False,
) -> tuple[pd.DataFrame, str]:
    """
    Reads ``xlsx_path`` (read-only), merges consecutive rows, and writes the
    result to a **new, separate file** in the same directory. The original
    workbook is never opened for writing or modified in any way.

    Returns (merged_dataframe, output_path).
    """
    if not os.path.isfile(xlsx_path):
        raise FileNotFoundError(f"Workbook not found: {xlsx_path}")

    df = pd.read_excel(xlsx_path, sheet_name=source_sheet)
    merged = merge_consecutive_activities(
        df,
        gap_seconds=gap_seconds,
        require_same_application=require_same_application,
    )

    out_dir = os.path.dirname(os.path.abspath(xlsx_path))
    out_path = os.path.join(out_dir, MERGED_FILE_NAME)
    sheet = "Merged Activities"
    merged.to_excel(out_path, index=False, sheet_name=sheet)
    if "Duration" in merged.columns:
        _format_duration_column(out_path, sheet, "Duration")

    return merged, out_path


def main(argv: list[str] | None = None) -> int:
    argv = argv if argv is not None else sys.argv[1:]
    path = argv[0] if argv else "All Activities.xlsx"
    gap = float(argv[1]) if len(argv) > 1 else DEFAULT_GAP_SECONDS

    merged, out_path = write_merged_file(path, gap_seconds=gap)
    print(f"Wrote {len(merged)} merged rows to: {out_path}")
    print(f"Original file untouched: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
